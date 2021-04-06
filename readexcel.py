# readexcel.py
# from dataexcel to web

from openpyxl import load_workbook
import time

execfile = load_workbook('fruit.xlsx')
sheet = execfile.active

# print(sheet['B1'].value) #การอ่าน header 

alltitle = []
allprice = []
alldata = []

for i in range(3,7):
	pdname = sheet.cell(row=i, column=2).value
	pdprice = sheet.cell(row=i, column=3).value
	pddetail = sheet.cell(row=i, column=4).value
	alltitle.append(pdname)
	allprice.append(pdprice)
	alldata.append(pddetail)


from selenium import webdriver
from selenium.webdriver.common.keys import Keys # เป็นคำสั่ง key 

driverpath = '/Users/ronnakorn/Desktop/Python Bootcamp 2021/EP.RPA with Python (Open by python 3.8)/chromedriver'
driver = webdriver.Chrome(driverpath)

url = 'http://www.uncle-machine.com/login/'

driver.get(url)


username = driver.find_element_by_id('username') # by_id or by_name ก้ได้ ขึ้นอยู่กับผู้พัฒนาเว็ปว่าจะใส่คำว่าอะไรมาให้ / ('') ในวงเล็บต้องตรงกับเว็ปตั้ง
username.send_keys("plug@gmail.com")			

password = driver.find_element_by_id('password') 
password.send_keys("0875058543")
# คือการสั่งให้ enter แทนการกดปุ่ม แต่ต้องimport key มาก่อน
password.send_keys(Keys.RETURN) 

# การส่งสำสั่งกดปุ่ม โดยวิธี xpath (ข้อเสียคือหารใช้ xpath เจ้าของเว็ปอาจเปลี่ยนแปลงคำสั่งได้ ทำให้โปรแกรมอาจerror)
# button = driver.find_element_by_xpath('/html/body/div[2]/form/button') 
# button.click()

addurl = 'http://www.uncle-machine.com/addproduct/'

driver.get(addurl) # go to addurl เพื่อเข้าไปในข้อมูลอื่นๆต่อไป

# ----------- Add Product ----------------------
for pn,pp,pd in zip(alltitle,allprice,alldata): # zip คือฟังชั่นการรวมข้อมูลในlistให้เป็นก้อนเดียวกัน 

	pdname = driver.find_element_by_id('name')
	pdprice = driver.find_element_by_id('price')
	pddetail = driver.find_element_by_id('detail')

	pd1_name = pn
	pd1_price = pp
	pd1_detail = pd


	pdname.send_keys(pd1_name)
	pdprice.send_keys(pd1_price)
	pddetail.send_keys(pd1_detail)
	time.sleep(1)

	button = driver.find_element_by_xpath('/html/body/div[2]/form/button') 
	button.click()

print('Upload ข้อมูลสำเร็จ')
